"""Run country profile optimizations without the GUI.

The runner configuration is intentionally hard-coded below. The workflow is:

1. Discover countries from the direct subfolders of ``COUNTRIES_ROOT``.
2. Load one base YAML into a ``ParameterObject``.
3. Iterate over ``SCENARIO_YEARS``.
4. Read and apply parameters from ``PARAMETERS_XLSX``.
5. Process one country at a time and its profiles in parallel.
6. Save a checkpoint Excel file after every completed country.

Run:
    python ptx_now/country_profile_runner.py
"""

from __future__ import annotations

import json
import multiprocessing
import os
import shutil
import tempfile
import time
import traceback
import zipfile
from copy import deepcopy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any


PROFILE_SUFFIXES = {".csv", ".xlsx", ".xls"}
RUNNER_VERSION = "2026-06-26-operational-balance-check-v4"
BALANCE_TOLERANCE = 1e-6
ZERO_CAPACITY_OUTPUT_SUM_TOLERANCE = 1e-3 * 8760
PROFILE_FEATURE_SHEET = "profile_features"
EXCEL_MAX_ROWS = 1_048_576


# ---------------------------------------------------------------------------
# Hard-coded configuration
# ---------------------------------------------------------------------------
COUNTRIES_ROOT = Path(
    r"/run/user/1000/gvfs/smb-share:server=iipsrv-ss1.iip.kit.edu,"
    r"share=daten$/weatherOut/weatherOut_Uwe"
)
SETTINGS_YAML = Path(
    r"/home/localadmin/Dokumente/ptx_now_data/"
    r"global_hydrogen_case_calculation/hydrogen.yaml"
)
PARAMETERS_XLSX = Path(
    r"/home/localadmin/Dokumente/ptx_now_data/"
    r"global_hydrogen_case_calculation/runner_parameters.xlsx"
)
WACC_FILE = Path(r"S:\Group_TE\GM_Uwe\transport_model\location_data_wacc.csv")
WACC_CSV = WACC_FILE  # Backwards-compatible alias for WACC_FILE.
OUTPUT_DIR = Path(
    r"/home/localadmin/Dokumente/ptx_now_data/"
    r"global_hydrogen_case_calculation/results"
)

SCENARIO_YEARS = (2030, 2040, 2050)
PROFILE_SUBDIR_TEMPLATE = "Clustered_Profiles/{year}/168"

SOLVER = "gurobi"
OPTIMIZATION_TYPE: str | None = "economical"
CORES: int | str = "max"
RECURSIVE_PROFILES = False
RUN_OPTIMIZATION = True
READ_PROFILE_STATISTICS = True
PROFILE_STATISTICS_CORES: int | str = 8
PROFILE_STATISTICS_RETRIES = 5
OPTIMIZATION_PROFILE_RETRIES = 5
SERVER_ACCESS_RETRIES: int | None = None
SERVER_ACCESS_RETRY_DELAY_SECONDS = 30.0

# None means all discovered country folders.
COUNTRIES: list[str] | None = None

COUNTRIES_SHEET = "countries"
PARAMETERS_SHEET = "parameters"
PROGRESS_FILE_NAME = "runner_progress.json"

ASSUMPTION_COMPONENT_LABELS = {
    "Solar": "Solar",
    "Wind": "Wind",
    "Electricity": "Battery",
    "electrolyzer": "Electrolyzer",
}
ASSUMPTION_PARAMETER_LABELS = {
    "capex": "CAPEX",
    "fixed_om": "FOM",
    "variable_om": "VOM",
    "charging_efficiency": "charge_eff",
    "discharging_efficiency": "discharge_eff",
    "ratio_capacity_p": "duration",
    "min_p": "min_p",
    "max_p": "max_p",
    "input.Electricity": "electricity_input",
}
ASSUMPTION_PARAMETER_ORDER = {
    parameter: order
    for order, parameter in enumerate(ASSUMPTION_PARAMETER_LABELS)
}

# Country folders are mapped to the parameter geographies below. Exact
# geographies such as China or India take precedence over broad regions.
EUROPE_COUNTRIES = {
    "Albania",
    "Austria",
    "Belarus",
    "Belgium",
    "Bosnia and Herzegovina",
    "Bulgaria",
    "Croatia",
    "Cyprus",
    "Czech Republic",
    "Denmark",
    "Estonia",
    "Finland",
    "France",
    "Germany",
    "Greenland",
    "Greece",
    "Hungary",
    "Iceland",
    "Ireland",
    "Italy",
    "Kosovo",
    "Latvia",
    "Lithuania",
    "Luxembourg",
    "Moldova",
    "Montenegro",
    "Netherlands",
    "North Macedonia",
    "Norway",
    "Poland",
    "Portugal",
    "Romania",
    "Serbia",
    "Slovakia",
    "Slovenia",
    "Spain",
    "Sweden",
    "Switzerland",
    "Turkey",
    "Ukraine",
    "United Kingdom",
}
MIDDLE_EAST_COUNTRIES = {
    "Bahrain",
    "Iran",
    "Iraq",
    "Israel",
    "Jordan",
    "Kuwait",
    "Lebanon",
    "Oman",
    "Palestine",
    "Palestinian Territory",
    "Qatar",
    "Saudi Arabia",
    "Syria",
    "United Arab Emirates",
    "Yemen",
}
AFRICA_COUNTRIES = {
    "Algeria",
    "Angola",
    "Benin",
    "Botswana",
    "Burkina Faso",
    "Burundi",
    "Cameroon",
    "Central African Republic",
    "Chad",
    "Comoros",
    "Democratic Republic of the Congo",
    "Djibouti",
    "Egypt",
    "Equatorial Guinea",
    "Eritrea",
    "Eswatini",
    "Ethiopia",
    "Gabon",
    "Gambia",
    "Ghana",
    "Guinea",
    "Guinea-Bissau",
    "Ivory Coast",
    "Kenya",
    "Lesotho",
    "Liberia",
    "Libya",
    "Madagascar",
    "Malawi",
    "Mali",
    "Mauritania",
    "Morocco",
    "Mozambique",
    "Namibia",
    "Niger",
    "Nigeria",
    "Republic of the Congo",
    "Rwanda",
    "Senegal",
    "Sierra Leone",
    "Somalia",
    "South Africa",
    "South Sudan",
    "Sudan",
    "Tanzania",
    "Togo",
    "Tunisia",
    "Uganda",
    "Western Sahara",
    "Zambia",
    "Zimbabwe",
}
LATIN_AMERICA_COUNTRIES = {
    "Argentina",
    "Belize",
    "Bolivia",
    "Chile",
    "Colombia",
    "Costa Rica",
    "Cuba",
    "Dominican Republic",
    "Ecuador",
    "El Salvador",
    "Guatemala",
    "Guyana",
    "Haiti",
    "Honduras",
    "Jamaica",
    "Mexico",
    "Nicaragua",
    "Panama",
    "Paraguay",
    "Peru",
    "Puerto Rico",
    "Suriname",
    "Uruguay",
    "Venezuela",
}
ASIA_OCEANIA_COUNTRIES = {
    "Afghanistan",
    "Armenia",
    "Australia",
    "Azerbaijan",
    "Bangladesh",
    "Bhutan",
    "Brunei",
    "Cambodia",
    "Georgia",
    "Indonesia",
    "Kazakhstan",
    "Kyrgyzstan",
    "Laos",
    "Malaysia",
    "Mongolia",
    "Myanmar",
    "Nepal",
    "New Zealand",
    "North Korea",
    "Pakistan",
    "Papua New Guinea",
    "Philippines",
    "South Korea",
    "Sri Lanka",
    "Tajikistan",
    "Taiwan",
    "Thailand",
    "Turkmenistan",
    "Uzbekistan",
    "Vietnam",
}
COUNTRY_REGION_ALIASES = {
    "Brazil": "Brazil",
    "Canada": "North America",
    "China": "China",
    "State of Palestine": "Middle East",
    "Palestinian Territories": "Middle East",
    "People's Republic of China": "China",
    "India": "India",
    "Japan": "Japan",
    "Russia": "Russia",
    "United States": "United States",
    "United States of America": "United States",
}


@dataclass(frozen=True)
class RunnerConfig:
    countries_root: Path
    settings_yaml: Path
    parameters_xlsx: Path
    wacc_file: Path | None
    output_dir: Path
    scenario_years: tuple[int, ...]
    profile_subdir_template: str
    solver: str
    optimization_type: str | None
    cores: int
    recursive_profiles: bool
    run_optimization: bool
    read_profile_statistics: bool
    profile_statistics_cores: int
    countries: list[str] | None


@dataclass(frozen=True)
class CountrySettings:
    country: str
    region: str
    wacc: float | None
    wacc_source: str


class OptimizationNotOptimalError(RuntimeError):
    def __init__(
        self,
        year: int,
        country: str,
        profile: str,
        reason: str,
        raw_status: str,
    ) -> None:
        self.year = year
        self.country = country
        self.profile = profile
        self.reason = reason
        self.raw_status = raw_status
        super().__init__(year, country, profile, reason, raw_status)

    def __str__(self) -> str:
        return (
            "Optimization aborted because no optimal solution was found: "
            f"year={self.year}, country={self.country}, "
            f"profile={self.profile}, reason={self.reason}, "
            f"solver_status={self.raw_status}"
        )


class ProfileOptimizationError(RuntimeError):
    def __init__(
        self,
        year: int,
        country: str,
        region: str,
        profile: str,
        profile_path: str,
        exception_type: str,
        detail: str,
        traceback_text: str,
    ) -> None:
        self.year = year
        self.country = country
        self.region = region
        self.profile = profile
        self.profile_path = profile_path
        self.exception_type = exception_type
        self.detail = detail
        self.traceback_text = traceback_text
        super().__init__(
            year,
            country,
            region,
            profile,
            profile_path,
            exception_type,
            detail,
            traceback_text,
        )

    def __str__(self) -> str:
        return (
            "Optimization aborted because the profile run failed: "
            f"year={self.year}, country={self.country}, "
            f"profile={self.profile}, profile_path={self.profile_path}, "
            f"exception={self.exception_type}, detail={self.detail}\n"
            f"Traceback:\n{self.traceback_text}"
        )


def _normalise_folder(path: Path) -> str:
    return str(path.resolve()) + os.sep


def _parse_cores(value: int | str) -> int:
    if isinstance(value, str) and value.lower() == "max":
        return max(1, multiprocessing.cpu_count() - 1)

    cores = int(value)
    if cores < 1:
        raise ValueError("CORES must be at least 1 or 'max'.")
    return cores


def build_config() -> RunnerConfig:
    return RunnerConfig(
        countries_root=COUNTRIES_ROOT,
        settings_yaml=SETTINGS_YAML,
        parameters_xlsx=PARAMETERS_XLSX,
        wacc_file=WACC_FILE or WACC_CSV,
        output_dir=OUTPUT_DIR,
        scenario_years=SCENARIO_YEARS,
        profile_subdir_template=PROFILE_SUBDIR_TEMPLATE,
        solver=SOLVER,
        optimization_type=OPTIMIZATION_TYPE,
        cores=_parse_cores(CORES),
        recursive_profiles=RECURSIVE_PROFILES,
        run_optimization=RUN_OPTIMIZATION,
        read_profile_statistics=READ_PROFILE_STATISTICS,
        profile_statistics_cores=_parse_cores(PROFILE_STATISTICS_CORES),
        countries=COUNTRIES,
    )


def validate_config(config: RunnerConfig) -> None:
    missing = []
    if not config.countries_root.is_dir():
        missing.append(f"COUNTRIES_ROOT does not exist: {config.countries_root}")
    if not config.settings_yaml.is_file():
        missing.append(f"SETTINGS_YAML does not exist: {config.settings_yaml}")
    if not config.parameters_xlsx.is_file():
        missing.append(f"PARAMETERS_XLSX does not exist: {config.parameters_xlsx}")
    if config.wacc_file is not None and not config.wacc_file.is_file():
        missing.append(f"WACC_FILE does not exist: {config.wacc_file}")
    if not config.run_optimization and not config.read_profile_statistics:
        missing.append(
            "At least one of RUN_OPTIMIZATION or READ_PROFILE_STATISTICS "
            "must be True."
        )

    if missing:
        details = "\n".join(f"- {message}" for message in missing)
        raise SystemExit(
            "Please correct the hard-coded configuration at the top of "
            f"country_profile_runner.py:\n{details}"
        )


def _load_case_data(settings_yaml: Path) -> dict[str, Any]:
    import yaml

    with settings_yaml.open("r", encoding="utf-8") as file:
        return yaml.load(file, Loader=yaml.FullLoader)


def _build_base_parameter_object(config: RunnerConfig) -> Any:
    from _load_projects import load_project
    from object_framework import ParameterObject

    case_data = _load_case_data(config.settings_yaml)
    pm_object = ParameterObject(
        "parameter",
        integer_steps=10,
        path_data=_normalise_folder(config.countries_root),
    )
    pm_object = load_project(pm_object, case_data)
    pm_object.set_solver(config.solver)
    if config.optimization_type:
        pm_object.set_optimization_type(config.optimization_type)
    return pm_object


def _is_active(value: Any) -> bool:
    if value is None:
        return True
    try:
        import pandas as pd

        if pd.isna(value):
            return True
    except TypeError:
        pass
    if isinstance(value, str):
        return value.strip().lower() not in {"false", "0", "no", "nein", "inactive"}
    return bool(value)


def _read_parameter_workbook(path: Path) -> tuple[Any, Any]:
    import pandas as pd

    countries = pd.read_excel(path, sheet_name=COUNTRIES_SHEET)
    parameter_matrix = pd.read_excel(path, sheet_name=PARAMETERS_SHEET)

    country_required = {"country", "region", "wacc"}
    parameter_required = {
        "country_or_region",
        "technology",
        "parameter",
        *[str(year) for year in SCENARIO_YEARS],
    }
    missing_country = country_required - set(countries.columns)
    parameter_matrix.columns = [
        str(column).strip() for column in parameter_matrix.columns
    ]
    missing_parameter = parameter_required - set(parameter_matrix.columns)
    if missing_country:
        raise ValueError(
            f"Sheet '{COUNTRIES_SHEET}' is missing columns: "
            f"{sorted(missing_country)}"
        )
    if missing_parameter:
        raise ValueError(
            f"Sheet '{PARAMETERS_SHEET}' is missing columns: "
            f"{sorted(missing_parameter)}"
        )

    if "active" in countries.columns:
        countries = countries[countries["active"].map(_is_active)]
    if "active" in parameter_matrix.columns:
        parameter_matrix = parameter_matrix[
            parameter_matrix["active"].map(_is_active)
        ]

    countries = countries.copy()
    parameter_matrix = parameter_matrix.copy()
    _fill_formula_parameter_values(path, parameter_matrix)
    countries["country"] = countries["country"].astype(str).str.strip()
    countries["region"] = countries["region"].astype(str).str.strip()

    parameter_matrix["country_or_region"] = (
        parameter_matrix["country_or_region"]
        .fillna("Global")
        .astype(str)
        .str.strip()
    )
    parameter_matrix["technology"] = (
        parameter_matrix["technology"].astype(str).str.strip()
    )
    parameter_matrix["parameter"] = (
        parameter_matrix["parameter"].astype(str).str.strip()
    )

    parameters = parameter_matrix.melt(
        id_vars=[
            column
            for column in parameter_matrix.columns
            if column not in {str(year) for year in SCENARIO_YEARS}
        ],
        value_vars=[str(year) for year in SCENARIO_YEARS],
        var_name="year",
        value_name="value",
    )
    parameters = parameters[pd.notna(parameters["value"])].copy()
    parameters["year"] = parameters["year"].astype(int)
    parameters["value"] = parameters["value"].astype(float)
    parameters["scope"] = "global"
    parameters["scope_name"] = ""
    geography = parameters["country_or_region"]
    country_names = set(countries["country"])
    region_names = set(countries["region"])
    region_mask = geography.isin(region_names)
    country_mask = geography.isin(country_names) & ~region_mask
    parameters.loc[region_mask, "scope"] = "region"
    parameters.loc[region_mask, "scope_name"] = parameters.loc[
        region_mask, "country_or_region"
    ]
    parameters.loc[country_mask, "scope"] = "country"
    parameters.loc[country_mask, "scope_name"] = parameters.loc[
        country_mask, "country_or_region"
    ]
    unknown_mask = (
        ~geography.str.casefold().eq("global")
        & ~region_mask
        & ~country_mask
    )
    if unknown_mask.any():
        unknown = sorted(set(geography[unknown_mask]))
        raise ValueError(
            "Unknown country_or_region entries in parameters sheet: "
            f"{unknown}. Add them to the countries sheet first."
        )
    parameters["component"] = parameters["technology"].map(
        _component_name_for_technology
    )
    parameters["parameter"] = parameters["parameter"].map(
        _internal_parameter_name
    )

    return countries, parameters


def _normalise_wacc(value: Any) -> float | None:
    import pandas as pd

    if value is None or pd.isna(value):
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        value = value.replace("%", "").strip()
        if "," in value and "." not in value:
            value = value.replace(",", ".")
    number = float(value)
    if abs(number) > 1:
        number = number / 100
    return number


def _find_column(columns: list[str], candidates: set[str]) -> str | None:
    import re

    def normalize(column: str) -> str:
        return re.sub(
            r"_+",
            "_",
            re.sub(r"[^a-z0-9]+", "_", column.strip().lower()),
        ).strip("_")

    normalized = {
        normalize(column): column
        for column in columns
    }
    for candidate in candidates:
        if candidate in normalized:
            return normalized[candidate]
    return None


def _read_wacc_file(path: Path | None) -> dict[tuple[str, int | None], float]:
    if path is None:
        return {}

    import pandas as pd

    suffix = path.suffix.lower()
    if suffix == ".csv":
        raw = pd.read_csv(path, sep=None, engine="python")
    elif suffix in {".xlsx", ".xls"}:
        raw = pd.read_excel(path)
    else:
        raise ValueError(
            f"Unsupported WACC file type '{path.suffix}'. "
            "Use .csv, .xlsx, or .xls."
        )
    raw.columns = [str(column).strip() for column in raw.columns]

    country_column = _find_column(
        list(raw.columns),
        {
            "country",
            "country_territory",
            "country_name",
            "country_or_region",
            "name",
        },
    )
    if country_column is None:
        raise ValueError(
            f"WACC file '{path}' needs a country column. Supported names are "
            "country, country_name, country_or_region, or name."
        )

    year_columns = [
        column
        for column in raw.columns
        if str(column).strip() in {str(year) for year in SCENARIO_YEARS}
    ]
    if year_columns:
        data = raw.melt(
            id_vars=[country_column],
            value_vars=year_columns,
            var_name="year",
            value_name="wacc",
        )
    else:
        wacc_column = _find_column(
            list(raw.columns),
            {
                "wacc",
                "weighted_average_cost_of_capital",
                "discount_rate",
                "value",
            },
        )
        if wacc_column is None:
            raise ValueError(
                f"WACC file '{path}' needs either year columns "
                f"{SCENARIO_YEARS} or a WACC column."
            )
        data = raw[[country_column, wacc_column]].copy()
        data = data.rename(columns={wacc_column: "wacc"})
        year_column = _find_column(list(raw.columns), {"year", "scenario_year"})
        if year_column is not None:
            data["year"] = raw[year_column]
        else:
            data["year"] = None

    data = data.rename(columns={country_column: "country"})
    data["country"] = data["country"].astype(str).str.strip()
    data = data[data["country"] != ""].copy()
    data["year"] = data["year"].map(
        lambda value: None if pd.isna(value) else int(value)
    )
    data["wacc"] = data["wacc"].map(_normalise_wacc)
    data = data.dropna(subset=["wacc"])

    duplicates = data.duplicated(subset=["country", "year"], keep=False)
    if duplicates.any():
        duplicated_rows = data.loc[duplicates, ["country", "year"]]
        raise ValueError(
            "WACC file has duplicate country/year rows: "
            f"{duplicated_rows.head(10).to_dict(orient='records')}"
        )

    return {
        (str(row["country"]), row["year"]): float(row["wacc"])
        for _, row in data.iterrows()
    }


def _fill_formula_parameter_values(path: Path, parameter_matrix: Any) -> None:
    import pandas as pd

    year_columns = [str(year) for year in SCENARIO_YEARS]
    if not any(
        column in parameter_matrix.columns
        and parameter_matrix[column].isna().any()
        for column in year_columns
    ):
        return

    try:
        import openpyxl
        from openpyxl.utils.cell import column_index_from_string
    except ImportError:
        return

    workbook = openpyxl.load_workbook(path, data_only=False, read_only=False)
    parameter_sheet = workbook[PARAMETERS_SHEET]
    header_to_column = {
        str(parameter_sheet.cell(1, column).value).strip(): column
        for column in range(1, parameter_sheet.max_column + 1)
    }
    cache: dict[tuple[str, str], float] = {}

    def split_cell_ref(reference: str, current_sheet: str) -> tuple[str, str]:
        reference = reference.replace("$", "").strip()
        if "!" not in reference:
            return current_sheet, reference
        sheet_name, cell_ref = reference.split("!", 1)
        return sheet_name.strip("'"), cell_ref

    def split_range_ref(reference: str, current_sheet: str) -> tuple[str, str, str]:
        reference = reference.replace("$", "").strip()
        if "!" in reference:
            sheet_name, range_ref = reference.split("!", 1)
            sheet_name = sheet_name.strip("'")
        else:
            sheet_name = current_sheet
            range_ref = reference
        start_ref, end_ref = range_ref.split(":", 1)
        return sheet_name, start_ref, end_ref

    def cell_value(sheet_name: str, cell_ref: str) -> float:
        cell_ref = cell_ref.replace("$", "")
        key = (sheet_name, cell_ref)
        if key in cache:
            return cache[key]
        value = workbook[sheet_name][cell_ref].value
        if isinstance(value, str) and value.startswith("="):
            value = evaluate_formula(value, sheet_name)
        if value is None or pd.isna(value):
            raise ValueError(f"Formula cell {sheet_name}!{cell_ref} is empty.")
        value = float(value)
        cache[key] = value
        return value

    def range_values(reference: str, current_sheet: str) -> list[float]:
        sheet_name, start_ref, end_ref = split_range_ref(reference, current_sheet)
        sheet = workbook[sheet_name]
        start_column = "".join(character for character in start_ref if character.isalpha())
        start_row = int("".join(character for character in start_ref if character.isdigit()))
        end_column = "".join(character for character in end_ref if character.isalpha())
        end_row = int("".join(character for character in end_ref if character.isdigit()))
        values = []
        for row in range(start_row, end_row + 1):
            for column in range(
                column_index_from_string(start_column),
                column_index_from_string(end_column) + 1,
            ):
                values.append(float(sheet.cell(row, column).value))
        return values

    def evaluate_index_match(formula: str, current_sheet: str) -> float:
        import re

        match = re.fullmatch(
            r"INDEX\((?P<value_range>.+),MATCH\((?P<lookup>.+),"
            r"(?P<lookup_range>.+),0\)\)",
            formula,
            flags=re.IGNORECASE,
        )
        if not match:
            raise ValueError(f"Unsupported INDEX/MATCH formula: ={formula}")
        value_range = range_values(match.group("value_range"), current_sheet)
        lookup_range = range_values(match.group("lookup_range"), current_sheet)
        lookup_sheet, lookup_ref = split_cell_ref(match.group("lookup"), current_sheet)
        lookup_value = cell_value(lookup_sheet, lookup_ref)
        for index, candidate in enumerate(lookup_range):
            if abs(candidate - lookup_value) <= 1e-9:
                return value_range[index]
        raise ValueError(f"No MATCH value found in formula: ={formula}")

    def evaluate_arithmetic(formula: str, current_sheet: str) -> float:
        import ast
        import operator
        import re

        operators = {
            ast.Add: operator.add,
            ast.Sub: operator.sub,
            ast.Mult: operator.mul,
            ast.Div: operator.truediv,
            ast.USub: operator.neg,
            ast.UAdd: operator.pos,
        }

        def replace_reference(match: Any) -> str:
            sheet_name = (match.group("quoted") or match.group("plain")).strip("'")
            cell_ref = match.group("cell").replace("$", "")
            return str(cell_value(sheet_name, cell_ref))

        expression = re.sub(
            r"(?P<sheet>(?:'(?P<quoted>[^']+)'|(?P<plain>[A-Za-z0-9_ &]+))!)"
            r"(?P<cell>\$?[A-Z]+\$?\d+)",
            replace_reference,
            formula,
        )

        def evaluate_node(node: Any) -> float:
            if isinstance(node, ast.Expression):
                return evaluate_node(node.body)
            if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
                return float(node.value)
            if isinstance(node, ast.BinOp) and type(node.op) in operators:
                return operators[type(node.op)](
                    evaluate_node(node.left),
                    evaluate_node(node.right),
                )
            if isinstance(node, ast.UnaryOp) and type(node.op) in operators:
                return operators[type(node.op)](evaluate_node(node.operand))
            raise ValueError(f"Unsupported arithmetic formula: ={formula}")

        return evaluate_node(ast.parse(expression, mode="eval"))

    def evaluate_formula(formula: str, current_sheet: str) -> float:
        import re

        formula = formula.strip()
        if formula.startswith("="):
            formula = formula[1:]
        if re.fullmatch(r"[+-]?\d+(\.\d+)?", formula):
            return float(formula)
        direct = re.fullmatch(
            r"(?:'(?P<quoted>[^']+)'|(?P<plain>[A-Za-z0-9_ &]+))!"
            r"(?P<cell>\$?[A-Z]+\$?\d+)",
            formula,
        )
        if direct:
            sheet_name = direct.group("quoted") or direct.group("plain")
            return cell_value(sheet_name, direct.group("cell"))
        if formula.upper().startswith("INDEX("):
            return evaluate_index_match(formula, current_sheet)
        return evaluate_arithmetic(formula, current_sheet)

    for row_index, row in parameter_matrix.iterrows():
        excel_row = int(row_index) + 2
        for year_column in year_columns:
            if year_column not in parameter_matrix.columns or not pd.isna(row[year_column]):
                continue
            excel_column = header_to_column.get(year_column)
            if excel_column is None:
                continue
            formula = parameter_sheet.cell(excel_row, excel_column).value
            if isinstance(formula, str) and formula.startswith("="):
                parameter_matrix.at[row_index, year_column] = evaluate_formula(
                    formula,
                    PARAMETERS_SHEET,
                )


def _component_name_for_technology(technology: str) -> str:
    mapping = {
        "solar": "Solar",
        "wind": "Wind",
        "battery": "Electricity",
        "electricity storage": "Electricity",
        "electrolyzer": "electrolyzer",
        "electrolyser": "electrolyzer",
    }
    return mapping.get(technology.strip().lower(), technology.strip())


def _internal_parameter_name(parameter: str) -> str:
    mapping = {
        "capex": "capex",
        "fom": "fixed_om",
        "fixed om": "fixed_om",
        "fixed_om": "fixed_om",
        "vom": "variable_om",
        "variable om": "variable_om",
        "variable_om": "variable_om",
        "charging efficiency": "charging_efficiency",
        "charging_efficiency": "charging_efficiency",
        "discharging efficiency": "discharging_efficiency",
        "discharging_efficiency": "discharging_efficiency",
        "duration": "ratio_capacity_p",
        "storage duration": "ratio_capacity_p",
        "ratio_capacity_p": "ratio_capacity_p",
        "minimum load": "min_p",
        "min_p": "min_p",
        "maximum load": "max_p",
        "max_p": "max_p",
        "electricity input": "input.Electricity",
        "electricity input per hydrogen": "input.Electricity",
        "electricity_input_per_hydrogen": "input.Electricity",
        "specific electricity consumption": "input.Electricity",
        "specific electricity consumption electricity": "input.Electricity",
        "mwh electricity per mwh hydrogen": "input.Electricity",
        "input.electricity": "input.Electricity",
    }
    return mapping.get(parameter.strip().lower(), parameter.strip())


def _discover_countries(config: RunnerConfig) -> list[Path]:
    country_dirs = [
        path for path in config.countries_root.iterdir() if path.is_dir()
    ]
    if config.countries:
        selected = set(config.countries)
        country_dirs = [path for path in country_dirs if path.name in selected]
    return sorted(country_dirs, key=lambda path: path.name.casefold())


def _country_settings(
    countries_df: Any,
    country: str,
    year: int,
    wacc_overrides: dict[tuple[str, int | None], float],
) -> CountrySettings:
    import pandas as pd

    matches = countries_df[countries_df["country"] == country]
    if len(matches) > 1:
        raise ValueError(
            f"Country '{country}' may occur at most once in sheet "
            f"'{COUNTRIES_SHEET}', found {len(matches)} rows."
        )

    available_regions = set(countries_df["region"])
    if len(matches) == 1:
        row = matches.iloc[0]
        region = str(row["region"]).strip()
        wacc = None if pd.isna(row["wacc"]) else float(row["wacc"])
        wacc_source = COUNTRIES_SHEET if wacc is not None else "base_yaml"
    else:
        region = _infer_parameter_region(country, available_regions)
        region_rows = countries_df[
            (countries_df["country"] == region)
            | (countries_df["region"] == region)
        ]
        region_wacc_values = region_rows["wacc"].dropna()
        wacc = (
            float(region_wacc_values.iloc[0])
            if len(region_wacc_values) > 0
            else None
        )
        wacc_source = (
            f"{COUNTRIES_SHEET}:{region}"
            if wacc is not None
            else "base_yaml"
        )

    if (country, year) in wacc_overrides:
        wacc = wacc_overrides[(country, year)]
        wacc_source = "wacc_csv"
    elif (country, None) in wacc_overrides:
        wacc = wacc_overrides[(country, None)]
        wacc_source = "wacc_csv"

    return CountrySettings(
        country=country,
        region=region,
        wacc=wacc,
        wacc_source=wacc_source,
    )


def _infer_parameter_region(
    country: str,
    available_regions: set[str],
) -> str:
    if country in COUNTRY_REGION_ALIASES:
        region = COUNTRY_REGION_ALIASES[country]
    elif country in EUROPE_COUNTRIES:
        region = "European Union"
    elif country in MIDDLE_EAST_COUNTRIES:
        region = "Middle East"
    elif country in AFRICA_COUNTRIES:
        region = "Africa"
    elif country in LATIN_AMERICA_COUNTRIES:
        region = "South & Latin America"
    elif country in ASIA_OCEANIA_COUNTRIES:
        region = "Asia & Oceania"
    else:
        raise ValueError(
            f"No automatic parameter region is defined for country "
            f"'{country}'. Add an exact row to sheet '{COUNTRIES_SHEET}' "
            "or extend the country mapping in country_profile_runner.py."
        )

    if region not in available_regions:
        raise ValueError(
            f"Country '{country}' maps to region '{region}', but that region "
            f"is not present in sheet '{COUNTRIES_SHEET}'."
        )
    return region


def _parameter_rows_for_country(
    parameters_df: Any,
    year: int,
    settings: CountrySettings,
) -> Any:
    year_rows = parameters_df[parameters_df["year"] == year].copy()
    selected = year_rows[
        (year_rows["scope"] == "global")
        | (
            (year_rows["scope"] == "region")
            & (year_rows["scope_name"] == settings.region)
        )
        | (
            (year_rows["scope"] == "country")
            & (year_rows["scope_name"] == settings.country)
        )
    ].copy()

    priority = {"global": 0, "region": 1, "country": 2}
    selected["_priority"] = selected["scope"].map(priority)
    selected = selected.sort_values(
        ["_priority", "component", "parameter"],
        kind="stable",
    )
    selected = selected.drop_duplicates(
        subset=["component", "parameter"],
        keep="last",
    )
    return selected


def _apply_parameter(pm_object: Any, component_name: str, parameter: str, value: float) -> None:
    if component_name == "__project__":
        if parameter != "wacc":
            raise ValueError(f"Unsupported project parameter: {parameter}")
        pm_object.set_wacc(value)
        return

    component = pm_object.get_component(component_name)
    setter_by_parameter = {
        "capex": component.set_capex,
        "fixed_om": component.set_fixed_OM,
        "variable_om": component.set_variable_OM,
    }

    if parameter in setter_by_parameter:
        setter_by_parameter[parameter](value)
    elif parameter == "charging_efficiency":
        component.set_charging_efficiency(value)
    elif parameter == "discharging_efficiency":
        component.set_discharging_efficiency(value)
    elif parameter == "ratio_capacity_p":
        component.set_ratio_capacity_p(value)
    elif parameter == "min_p":
        component.set_min_p(value)
    elif parameter == "max_p":
        component.set_max_p(value)
    elif parameter.startswith("input."):
        commodity = parameter.split(".", 1)[1]
        component.add_input(commodity, value)
    elif parameter.startswith("output."):
        commodity = parameter.split(".", 1)[1]
        component.add_output(commodity, value)
    else:
        raise ValueError(
            f"Unsupported parameter '{parameter}' for component "
            f"'{component_name}'."
        )


def _apply_parameters(pm_object: Any, parameter_rows: Any) -> list[dict[str, Any]]:
    applied = []
    for _, row in parameter_rows.iterrows():
        _apply_parameter(
            pm_object,
            component_name=row["component"],
            parameter=row["parameter"],
            value=float(row["value"]),
        )
        applied.append(
            {
                key: row.get(key)
                for key in [
                    "year",
                    "scope",
                    "scope_name",
                    "component",
                    "parameter",
                    "value",
                    "unit",
                    "source",
                    "note",
                ]
                if key in row.index
            }
        )
    return applied


def _format_assumption_value(value: Any) -> str:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    if number != number:
        return "n/a"
    if number == 0:
        return "0"
    if abs(number) >= 100:
        return f"{number:.2f}"
    return f"{number:.6g}"


def _format_wacc(value: Any) -> str:
    try:
        return f"{float(value):.2%}"
    except (TypeError, ValueError):
        return "n/a"


def _print_country_assumptions(
    year: int,
    settings: CountrySettings,
    parameter_rows: Any,
    pm_object: Any,
) -> None:
    try:
        wacc = pm_object.get_wacc()
    except AttributeError:
        wacc = settings.wacc

    print(
        f"Assumptions {year}/{settings.country} "
        f"({settings.region}): WACC={_format_wacc(wacc)}"
    )

    central_rows = parameter_rows[
        parameter_rows["component"].isin(ASSUMPTION_COMPONENT_LABELS)
        & parameter_rows["parameter"].isin(ASSUMPTION_PARAMETER_LABELS)
    ].copy()
    if central_rows.empty:
        print("  No central CAPEX/operation assumptions found.")
        return

    central_rows["_component_order"] = central_rows["component"].map(
        {
            component: order
            for order, component in enumerate(ASSUMPTION_COMPONENT_LABELS)
        }
    )
    central_rows["_parameter_order"] = central_rows["parameter"].map(
        ASSUMPTION_PARAMETER_ORDER
    )
    central_rows = central_rows.sort_values(
        ["_component_order", "_parameter_order"],
        kind="stable",
    )

    for component, rows in central_rows.groupby("component", sort=False):
        values = []
        for _, row in rows.iterrows():
            label = ASSUMPTION_PARAMETER_LABELS[row["parameter"]]
            value = _format_assumption_value(row["value"])
            values.append(f"{label}={value}")
        print(f"  {ASSUMPTION_COMPONENT_LABELS[component]}: {', '.join(values)}")


def _profile_dir(config: RunnerConfig, country_dir: Path, year: int) -> Path:
    relative = config.profile_subdir_template.format(year=year)
    return country_dir / Path(relative)


def _is_retryable_server_access_error(exc: BaseException) -> bool:
    if isinstance(exc, (EOFError, zipfile.BadZipFile)):
        return True
    message = str(exc).lower()
    if (
        "io.excel.zip.reader" in message
        or "no such keys" in message
        or "no such key" in message
        or "file is not a zip file" in message
        or "badzipfile" in message
    ):
        return True
    if isinstance(exc, ValueError):
        return (
            "excel file format cannot be determined" in message
        )
    if not isinstance(exc, OSError):
        return False
    retryable_error_numbers = {
        2,    # File or directory temporarily invisible on flaky mounts.
        5,    # Input/output error.
        22,   # Invalid argument while streaming xlsx from GVFS/SMB.
        107,  # Transport endpoint is not connected.
        110,  # Connection timed out.
        112,  # Host is down.
        113,  # No route to host.
        121,  # Remote I/O error.
    }
    retryable_windows_errors = {
        53,    # Network path not found.
        59,    # Unexpected network error.
        64,    # Network name deleted.
        121,   # Semaphore timeout period expired.
        1231,  # Network location cannot be reached.
    }
    return (
        getattr(exc, "errno", None) in retryable_error_numbers
        or getattr(exc, "winerror", None) in retryable_windows_errors
    )


def _optimization_profile_retryable(exc: BaseException) -> bool:
    message = str(exc).lower()
    return (
        _is_retryable_server_access_error(exc)
        or "file is not a zip file" in message
        or "badzipfile" in message
        or "excel file format cannot be determined" in message
        or "could not create a valid local copy" in message
    )


def _retry_server_access(description: str, operation: Any) -> Any:
    attempt = 0
    while True:
        try:
            return operation()
        except Exception as exc:  # noqa: BLE001 - transient server access.
            if not _is_retryable_server_access_error(exc):
                raise

            attempt += 1
            if (
                SERVER_ACCESS_RETRIES is not None
                and attempt > SERVER_ACCESS_RETRIES
            ):
                raise
            time.sleep(SERVER_ACCESS_RETRY_DELAY_SECONDS)


def _print_progress(label: str, completed: int, total: int) -> None:
    if total <= 0:
        return
    width = 30
    fraction = min(1.0, max(0.0, completed / total))
    filled = int(round(width * fraction))
    bar = "#" * filled + "-" * (width - filled)
    percent = 100 * fraction
    print(
        f"\r{label}: [{bar}] {completed}/{total} ({percent:5.1f}%)",
        end="",
        flush=True,
    )


def _finish_progress(label: str, completed: int, total: int) -> None:
    _print_progress(label, completed, total)
    print()


def _profile_files(profile_dir: Path, recursive: bool) -> list[str]:
    def list_profiles() -> list[str]:
        if recursive:
            files = [
                path
                for path in profile_dir.rglob("*")
                if path.is_file() and path.suffix.lower() in PROFILE_SUFFIXES
            ]
            return sorted(str(path.relative_to(profile_dir)) for path in files)

        return sorted(
            path.name
            for path in profile_dir.iterdir()
            if path.is_file() and path.suffix.lower() in PROFILE_SUFFIXES
        )

    return _retry_server_access(
        f"profile listing {profile_dir}",
        list_profiles,
    )


def _wait_for_profile_dir(profile_dir: Path) -> None:
    def check_profile_dir() -> None:
        if not profile_dir.is_dir():
            raise FileNotFoundError(
                f"Profile directory not available: {profile_dir}"
            )

    _retry_server_access(
        f"profile directory {profile_dir}",
        check_profile_dir,
    )


def _read_profile_frame(path: Path) -> Any:
    import pandas as pd

    if path.suffix.lower() in {".xlsx", ".xls"}:
        return pd.read_excel(path, index_col=0)
    return pd.read_csv(path, index_col=0)


def _profile_column(frame: Any, column_name: str) -> str | None:
    normalized = {
        str(column).strip().lower(): column
        for column in frame.columns
    }
    return normalized.get(column_name.strip().lower())


def _weighted_average(values: Any, weights: Any) -> float | None:
    import numpy as np

    valid = np.isfinite(values) & np.isfinite(weights) & (weights > 0)
    if not valid.any():
        return None
    return float(np.average(values[valid], weights=weights[valid]))


def _weighted_std(values: Any, weights: Any) -> float | None:
    import numpy as np

    mean = _weighted_average(values, weights)
    if mean is None:
        return None
    valid = np.isfinite(values) & np.isfinite(weights) & (weights > 0)
    variance = np.average((values[valid] - mean) ** 2, weights=weights[valid])
    return float(np.sqrt(variance))


def _weighted_share_below(values: Any, weights: Any, threshold: float) -> float | None:
    import numpy as np

    valid = np.isfinite(values) & np.isfinite(weights) & (weights > 0)
    if not valid.any():
        return None
    total_weight = float(weights[valid].sum())
    if total_weight <= 0:
        return None
    return float(weights[valid & (values < threshold)].sum() / total_weight)


def _weighted_correlation(
    left_values: Any,
    right_values: Any,
    weights: Any,
) -> float | None:
    import numpy as np

    valid = (
        np.isfinite(left_values)
        & np.isfinite(right_values)
        & np.isfinite(weights)
        & (weights > 0)
    )
    if not valid.any():
        return None
    left = left_values[valid]
    right = right_values[valid]
    valid_weights = weights[valid]
    left_mean = np.average(left, weights=valid_weights)
    right_mean = np.average(right, weights=valid_weights)
    covariance = np.average(
        (left - left_mean) * (right - right_mean),
        weights=valid_weights,
    )
    left_std = np.sqrt(np.average((left - left_mean) ** 2, weights=valid_weights))
    right_std = np.sqrt(
        np.average((right - right_mean) ** 2, weights=valid_weights)
    )
    if left_std <= 0 or right_std <= 0:
        return None
    return float(covariance / (left_std * right_std))


def _profile_series_features(
    frame: Any,
    column_name: str,
    weights: Any,
    prefix: str,
) -> dict[str, Any]:
    import numpy as np

    column = _profile_column(frame, column_name)
    if column is None:
        return {}

    values = frame[column].to_numpy(dtype=float)
    valid = np.isfinite(values) & np.isfinite(weights) & (weights > 0)
    if not valid.any():
        return {
            f"{prefix}_has_profile": False,
        }

    weighted_mean = _weighted_average(values, weights)
    weighted_std = _weighted_std(values, weights)
    return {
        f"{prefix}_has_profile": True,
        f"{prefix}_full_load_hours": float((values[valid] * weights[valid]).sum()),
        f"{prefix}_weighted_mean": weighted_mean,
        f"{prefix}_weighted_std": weighted_std,
        f"{prefix}_coefficient_of_variation": (
            weighted_std / weighted_mean
            if weighted_mean not in {None, 0}
            else None
        ),
        f"{prefix}_min": float(np.min(values[valid])),
        f"{prefix}_p10": float(np.percentile(values[valid], 10)),
        f"{prefix}_p50": float(np.percentile(values[valid], 50)),
        f"{prefix}_p90": float(np.percentile(values[valid], 90)),
        f"{prefix}_max": float(np.max(values[valid])),
        f"{prefix}_share_below_0_05": _weighted_share_below(
            values,
            weights,
            0.05,
        ),
        f"{prefix}_share_below_0_10": _weighted_share_below(
            values,
            weights,
            0.10,
        ),
    }


def calculate_profile_features(
    profile_path: Path,
    *,
    year: int,
    country: str,
    region: str,
    profile: str,
) -> dict[str, Any]:
    import numpy as np

    frame = _read_profile_frame(profile_path)
    weighting_column = _profile_column(frame, "Weighting")
    if weighting_column is None:
        weights = np.ones(len(frame), dtype=float)
    else:
        weights = frame[weighting_column].to_numpy(dtype=float)

    record = {
        "scenario_year": year,
        "country": country,
        "region": region,
        "profile": profile,
        "profile_rows": len(frame),
        "total_weighted_hours": float(np.nansum(weights)),
        "weighting_column_present": weighting_column is not None,
    }
    record.update(_profile_series_features(frame, "Solar", weights, "solar"))
    record.update(_profile_series_features(frame, "Wind", weights, "wind"))

    solar_column = _profile_column(frame, "Solar")
    wind_column = _profile_column(frame, "Wind")
    if solar_column is not None and wind_column is not None:
        solar = frame[solar_column].to_numpy(dtype=float)
        wind = frame[wind_column].to_numpy(dtype=float)
        valid = (
            np.isfinite(solar)
            & np.isfinite(wind)
            & np.isfinite(weights)
            & (weights > 0)
        )
        if valid.any():
            total_weight = float(weights[valid].sum())
            record["solar_wind_correlation"] = _weighted_correlation(
                solar,
                wind,
                weights,
            )
            record["solar_wind_weighted_mean_sum"] = _weighted_average(
                solar + wind,
                weights,
            )
            record["share_solar_and_wind_below_0_10"] = (
                float(
                    weights[
                        valid
                        & (solar < 0.10)
                        & (wind < 0.10)
                    ].sum()
                    / total_weight
                )
                if total_weight > 0
                else None
            )

    return record


def _profile_feature_records(
    profile_dir: Path,
    profiles: list[str],
    *,
    year: int,
    country: str,
    region: str,
    workers: int,
) -> list[dict[str, Any]]:
    records = []
    label = f"{year} {country} profile statistics"
    total = len(profiles)
    _print_progress(label, 0, total)

    jobs = [
        {
            "profile_dir": profile_dir,
            "profile": profile,
            "year": year,
            "country": country,
            "region": region,
        }
        for profile in profiles
    ]
    worker_count = min(max(1, workers), total)
    if worker_count == 1:
        for completed, job in enumerate(jobs, start=1):
            records.append(_calculate_profile_features_job(job))
            _print_progress(label, completed, total)
        _finish_progress(label, len(records), total)
        return records

    with multiprocessing.Pool(
        processes=worker_count,
        maxtasksperchild=20,
    ) as pool:
        for completed, record in enumerate(
            pool.imap_unordered(_calculate_profile_features_job, jobs),
            start=1,
        ):
            records.append(record)
            _print_progress(label, completed, total)

    _finish_progress(label, len(records), total)
    return records


def _calculate_profile_features_job(job: dict[str, Any]) -> dict[str, Any]:
    profile_dir = Path(job["profile_dir"])
    profile = job["profile"]
    profile_path = profile_dir / profile
    safe_country = "".join(
        character if character.isalnum() else "_"
        for character in str(job["country"])
    )

    last_error: BaseException | None = None
    for attempt in range(1, PROFILE_STATISTICS_RETRIES + 1):
        try:
            record = calculate_profile_features(
                profile_path,
                year=job["year"],
                country=job["country"],
                region=job["region"],
                profile=profile,
            )
            record["profile_feature_status"] = "ok"
            record["profile_feature_error"] = None
            return record
        except Exception as exc:  # noqa: BLE001 - flaky profile reads use local retry.
            last_error = exc

        local_profile_dir = Path(
            tempfile.mkdtemp(
                prefix=f"ptx_now_stats_{job['year']}_{safe_country}_",
            )
        )
        local_profile = local_profile_dir / profile
        try:
            _copy_profile_with_retries(
                source=profile_path,
                destination=local_profile,
                max_attempts=1,
            )
            record = calculate_profile_features(
                local_profile,
                year=job["year"],
                country=job["country"],
                region=job["region"],
                profile=profile,
            )
            record["profile_feature_status"] = "ok"
            record["profile_feature_error"] = None
            return record
        except Exception as exc:  # noqa: BLE001 - retry flaky local copy reads.
            last_error = exc
        finally:
            shutil.rmtree(local_profile_dir, ignore_errors=True)

        if attempt < PROFILE_STATISTICS_RETRIES:
            time.sleep(SERVER_ACCESS_RETRY_DELAY_SECONDS)

    return {
        "scenario_year": job["year"],
        "country": job["country"],
        "region": job["region"],
        "profile": profile,
        "profile_rows": None,
        "total_weighted_hours": None,
        "weighting_column_present": None,
        "profile_feature_status": "failed",
        "profile_feature_error": repr(last_error),
    }


def _validate_staged_profile(path: Path) -> None:
    if not path.is_file():
        raise FileNotFoundError(f"Staged profile does not exist: {path}")
    if path.stat().st_size == 0:
        raise OSError(f"Staged profile is empty: {path}")

    if path.suffix.lower() == ".xlsx":
        with zipfile.ZipFile(path, "r") as archive:
            corrupt_member = archive.testzip()
            if corrupt_member is not None:
                raise zipfile.BadZipFile(
                    f"Corrupt member '{corrupt_member}' in {path}"
                )


def _copy_profile_with_retries(
    source: Path,
    destination: Path,
    max_attempts: int | None = None,
) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary_destination = destination.with_suffix(
        destination.suffix + ".part"
    )
    last_error: Exception | None = None

    attempt = 0
    while True:
        attempt += 1
        try:
            temporary_destination.unlink(missing_ok=True)
            destination.unlink(missing_ok=True)

            source_size_before = source.stat().st_size
            with source.open("rb") as source_file:
                with temporary_destination.open("wb") as destination_file:
                    shutil.copyfileobj(
                        source_file,
                        destination_file,
                        length=1024 * 1024,
                    )
                    destination_file.flush()
                    os.fsync(destination_file.fileno())

            source_size_after = source.stat().st_size
            copied_size = temporary_destination.stat().st_size
            if source_size_before != source_size_after:
                raise OSError(
                    "Source file size changed while it was being copied: "
                    f"{source_size_before} -> {source_size_after}"
                )
            if copied_size != source_size_after:
                raise OSError(
                    f"Incomplete copy: source={source_size_after} bytes, "
                    f"local={copied_size} bytes"
                )

            _validate_staged_profile(temporary_destination)
            os.replace(temporary_destination, destination)
            return

        except Exception as exc:  # noqa: BLE001 - network reads are retried.
            last_error = exc
            temporary_destination.unlink(missing_ok=True)
            destination.unlink(missing_ok=True)
            if (
                max_attempts is not None
                and attempt >= max_attempts
            ) or (
                SERVER_ACCESS_RETRIES is not None
                and attempt >= SERVER_ACCESS_RETRIES
            ):
                break

            # The source profiles live on a flaky SMB/GVFS mount. Even errors
            # that are not neatly classified by Python often disappear on the
            # next read, so copying keeps retrying until the configured limit.
            time.sleep(SERVER_ACCESS_RETRY_DELAY_SECONDS)

    raise OSError(
        f"Could not create a valid local copy after "
        f"{attempt} attempts: {source}; last_error={last_error!r}"
    ) from last_error


def _model_class_for_solver(solver: str):
    if solver.lower() == "gurobi":
        from optimization_gurobi_model import OptimizationGurobiModel

        return OptimizationGurobiModel

    from optimization_pyomo_model import OptimizationPyomoModel

    return OptimizationPyomoModel


def _suppress_solver_log(optimization_problem: Any, solver: str) -> None:
    if solver.lower() != "gurobi":
        return
    model = getattr(optimization_problem, "model", None)
    if model is None:
        return

    for parameter, value in [
        ("OutputFlag", 0),
        ("LogToConsole", 0),
    ]:
        try:
            model.setParam(parameter, value)
        except Exception:
            pass


def _optimization_status(
    optimization_problem: Any,
) -> tuple[bool, str, str]:
    status = getattr(optimization_problem, "status", None)
    if status is not None:
        gurobi_status = {
            1: "loaded",
            2: "optimal",
            3: "infeasible",
            4: "infeasible_or_unbounded",
            5: "unbounded",
            6: "cutoff",
            7: "iteration_limit",
            8: "node_limit",
            9: "time_limit",
            10: "solution_limit",
            11: "interrupted",
            12: "numeric_error",
            13: "suboptimal",
            14: "in_progress",
            15: "user_objective_limit",
            16: "work_limit",
            17: "memory_limit",
        }
        reason = gurobi_status.get(int(status), "unknown")
        return status == 2, reason, f"gurobi_status_{status}"

    results = getattr(optimization_problem, "results", None)
    if results is None:
        return False, "unknown", "unknown"

    solver_status = str(results.solver.status).lower()
    termination = str(results.solver.termination_condition).lower()
    normalized_termination = termination.replace(" ", "_")
    if "infeasibleorunbounded" in normalized_termination.replace("_", ""):
        reason = "infeasible_or_unbounded"
    elif "infeasible" in normalized_termination:
        reason = "infeasible"
    elif "unbounded" in normalized_termination:
        reason = "unbounded"
    elif "timelimit" in normalized_termination.replace("_", ""):
        reason = "time_limit"
    elif normalized_termination == "optimal":
        reason = "optimal"
    else:
        reason = normalized_termination

    return (
        solver_status == "ok" and termination == "optimal",
        reason,
        f"{solver_status}_{termination}",
    )


def _variable_value(variable: Any) -> float:
    value = getattr(variable, "X", None)
    if value is None:
        value = getattr(variable, "value", None)
    if callable(value):
        value = value()
    if value is None:
        raise ValueError(f"Solver variable has no result value: {variable!r}")
    return float(value)


def _solver_nominal_capacity(
    optimization_problem: Any,
    component_name: str,
) -> float:
    nominal_cap = getattr(optimization_problem, "nominal_cap", None)
    if nominal_cap is None:
        instance = getattr(optimization_problem, "instance", None)
        nominal_cap = getattr(instance, "nominal_cap", None)
    if nominal_cap is None:
        raise ValueError("Optimization result has no nominal_cap variable.")
    return _variable_value(nominal_cap[component_name])


def _solver_weighted_flow_sum(
    variable: Any,
    keys: list[tuple[Any, ...]],
    weightings: dict[Any, float],
    cluster_index: int,
) -> float:
    quantity = 0.0
    for key in keys:
        quantity += _variable_value(variable[key]) * weightings[key[cluster_index]]
    return quantity


def _solver_variable(optimization_problem: Any, name: str) -> Any:
    variable = getattr(optimization_problem, name, None)
    if variable is None:
        instance = getattr(optimization_problem, "instance", None)
        variable = getattr(instance, name, None)
    return variable


def _solver_value_at(
    optimization_problem: Any,
    variable_name: str,
    key: tuple[Any, ...],
) -> float:
    variable = _solver_variable(optimization_problem, variable_name)
    if variable is None:
        return 0.0
    try:
        return _variable_value(variable[key])
    except (KeyError, IndexError, TypeError):
        return 0.0


def _solver_component_output_quantity(
    optimization_problem: Any,
    pm_object: Any,
    component_name: str,
    commodity_names: set[str],
) -> float:
    variable = _solver_variable(
        optimization_problem,
        "mass_energy_component_out_commodities",
    )
    if variable is None:
        return 0.0

    weightings = pm_object.get_weightings_time_series()
    keys = [
        (component_name, commodity_name, cl, t)
        for commodity_name in commodity_names
        for cl in range(pm_object.get_number_clusters())
        for t in range(pm_object.get_covered_period())
    ]
    existing_keys = [key for key in keys if key in variable]
    return _solver_weighted_flow_sum(
        variable,
        existing_keys,
        weightings,
        cluster_index=2,
    )


def _solver_generator_output_quantity(
    optimization_problem: Any,
    pm_object: Any,
    component_name: str,
    commodity_name: str,
) -> float:
    variable = _solver_variable(optimization_problem, "mass_energy_generation")
    if variable is None:
        return 0.0

    weightings = pm_object.get_weightings_time_series()
    keys = [
        (component_name, commodity_name, cl, t)
        for cl in range(pm_object.get_number_clusters())
        for t in range(pm_object.get_covered_period())
    ]
    existing_keys = [key for key in keys if key in variable]
    return _solver_weighted_flow_sum(
        variable,
        existing_keys,
        weightings,
        cluster_index=2,
    )


def _format_balance_terms(terms: list[tuple[str, float]]) -> str:
    relevant_terms = [
        (name, value)
        for name, value in terms
        if abs(value) > BALANCE_TOLERANCE
    ]
    if not relevant_terms:
        return "none"
    relevant_terms.sort(key=lambda item: abs(item[1]), reverse=True)
    return ", ".join(
        f"{name}={value:.12g}"
        for name, value in relevant_terms[:12]
    )


def _solver_operational_flow_records(
    optimization_problem: Any,
    pm_object: Any,
    year: int,
    country: str,
    region: str,
    profile: str,
) -> list[dict[str, Any]]:
    conversion_components = pm_object.get_final_conversion_components_objects()
    generator_components = pm_object.get_final_generator_components_objects()
    storage_names = set(pm_object.get_final_storage_components_names())
    commodity_objects = pm_object.get_final_commodities_objects()
    weightings = pm_object.get_weightings_time_series()
    records = []

    for commodity in commodity_objects:
        commodity_name = commodity.get_name()
        totals = {
            "available": 0.0,
            "purchased": 0.0,
            "generated": 0.0,
            "conversion_output": 0.0,
            "storage_out": 0.0,
            "emitted": 0.0,
            "sold": 0.0,
            "demanded": 0.0,
            "conversion_input": 0.0,
            "storage_in": 0.0,
            "hot_standby_demand": 0.0,
        }

        for cl in range(pm_object.get_number_clusters()):
            weighting = weightings[cl]
            for t in range(pm_object.get_covered_period()):
                source_terms = []
                sink_terms = []

                if commodity.is_available():
                    value = _solver_value_at(
                        optimization_problem,
                        "mass_energy_available",
                        (commodity_name, cl, t),
                    )
                    source_terms.append(("available", value))
                    totals["available"] += value * weighting
                if commodity.is_purchasable():
                    value = _solver_value_at(
                        optimization_problem,
                        "mass_energy_purchase_commodity",
                        (commodity_name, cl, t),
                    )
                    source_terms.append(("purchased", value))
                    totals["purchased"] += value * weighting
                if commodity.is_emittable():
                    value = _solver_value_at(
                        optimization_problem,
                        "mass_energy_emitted",
                        (commodity_name, cl, t),
                    )
                    sink_terms.append(("emitted", value))
                    totals["emitted"] += value * weighting
                if commodity.is_saleable():
                    value = _solver_value_at(
                        optimization_problem,
                        "mass_energy_sell_commodity",
                        (commodity_name, cl, t),
                    )
                    sink_terms.append(("sold", value))
                    totals["sold"] += value * weighting
                if commodity.is_demanded():
                    value = _solver_value_at(
                        optimization_problem,
                        "mass_energy_demand",
                        (commodity_name, cl, t),
                    )
                    sink_terms.append(("demanded", value))
                    totals["demanded"] += value * weighting

                if commodity_name in storage_names:
                    value = _solver_value_at(
                        optimization_problem,
                        "mass_energy_storage_out_commodities",
                        (commodity_name, cl, t),
                    )
                    source_terms.append(("storage_out", value))
                    totals["storage_out"] += value * weighting

                    value = _solver_value_at(
                        optimization_problem,
                        "mass_energy_storage_in_commodities",
                        (commodity_name, cl, t),
                    )
                    sink_terms.append(("storage_in", value))
                    totals["storage_in"] += value * weighting

                for generator in generator_components:
                    if generator.get_generated_commodity() != commodity_name:
                        continue
                    value = _solver_value_at(
                        optimization_problem,
                        "mass_energy_generation",
                        (generator.get_name(), commodity_name, cl, t),
                    )
                    source_terms.append((f"generated:{generator.get_name()}", value))
                    totals["generated"] += value * weighting

                for component in conversion_components:
                    component_name = component.get_name()
                    if commodity_name in component.get_outputs():
                        value = _solver_value_at(
                            optimization_problem,
                            "mass_energy_component_out_commodities",
                            (component_name, commodity_name, cl, t),
                        )
                        source_terms.append(
                            (f"conversion_output:{component_name}", value)
                        )
                        totals["conversion_output"] += value * weighting
                    if commodity_name in component.get_inputs():
                        value = _solver_value_at(
                            optimization_problem,
                            "mass_energy_component_in_commodities",
                            (component_name, commodity_name, cl, t),
                        )
                        sink_terms.append(
                            (f"conversion_input:{component_name}", value)
                        )
                        totals["conversion_input"] += value * weighting
                    hot_standby = component.get_hot_standby_demand()
                    if commodity_name in hot_standby:
                        value = _solver_value_at(
                            optimization_problem,
                            "mass_energy_hot_standby_demand",
                            (component_name, commodity_name, cl, t),
                        )
                        sink_terms.append(
                            (f"hot_standby_demand:{component_name}", value)
                        )
                        totals["hot_standby_demand"] += value * weighting

                source_total = sum(value for _, value in source_terms)
                sink_total = sum(value for _, value in sink_terms)
                residual = source_total - sink_total
                scale = max(abs(source_total), abs(sink_total), 1.0)
                if abs(residual) > max(BALANCE_TOLERANCE, BALANCE_TOLERANCE * scale):
                    raise RuntimeError(
                        "Operational commodity balance check failed: "
                        f"commodity={commodity_name}, cluster={cl}, time={t}, "
                        f"sources={source_total:.12g}, sinks={sink_total:.12g}, "
                        f"residual={residual:.12g}, "
                        f"source_terms=[{_format_balance_terms(source_terms)}], "
                        f"sink_terms=[{_format_balance_terms(sink_terms)}]"
                    )

        source_total = (
            totals["available"]
            + totals["purchased"]
            + totals["generated"]
            + totals["conversion_output"]
            + totals["storage_out"]
        )
        sink_total = (
            totals["emitted"]
            + totals["sold"]
            + totals["demanded"]
            + totals["conversion_input"]
            + totals["storage_in"]
            + totals["hot_standby_demand"]
        )
        records.append(
            {
                "scenario_year": year,
                "country": country,
                "region": region,
                "profile": profile,
                "commodity": commodity_name,
                **totals,
                "source_total": source_total,
                "sink_total": sink_total,
                "balance_residual": source_total - sink_total,
            }
        )

    return records


def _run_single_profile(job: dict[str, Any]) -> dict[str, Any]:
    country = job["country"]
    region = job["region"]
    year = job["year"]
    profile = job["profile"]

    try:
        from _helper_optimization import clone_components_which_use_parallelization
        from _transfer_results_to_parameter_object import (
            _transfer_results_to_parameter_object,
        )

        pm_object = clone_components_which_use_parallelization(job["pm_object"])
        pm_object.set_path_data(_normalise_folder(job["profile_dir"]))
        pm_object.set_profile_data(profile)
        pm_object.set_solver(job["solver"])

        optimization_model = _model_class_for_solver(job["solver"])
        optimization_problem = optimization_model(pm_object, job["solver"])
        _suppress_solver_log(optimization_problem, job["solver"])
        optimization_problem.prepare(
            optimization_type=pm_object.get_optimization_type()
        )
        try:
            optimization_problem.optimize()
        except Exception as exc:
            is_optimal, reason, raw_status = _optimization_status(
                optimization_problem
            )
            if not is_optimal and raw_status != "unknown":
                raise OptimizationNotOptimalError(
                    year,
                    country,
                    profile,
                    reason,
                    raw_status,
                ) from exc
            raise

        is_optimal, reason, solver_status = _optimization_status(
            optimization_problem
        )
        if not is_optimal:
            raise OptimizationNotOptimalError(
                year,
                country,
                profile,
                reason,
                solver_status,
            )

        economic_objective = getattr(
            optimization_problem,
            "economic_objective_function_value",
            getattr(optimization_problem, "objective_function_value", None),
        )
        ecologic_objective = getattr(
            optimization_problem,
            "ecologic_objective_function_value",
            None,
        )

        base_result = {
            "scenario_year": year,
            "country": country,
            "region": region,
            "profile": profile,
            "wacc": job["wacc"],
            "solver_status": solver_status,
            "economic_objective": economic_objective,
            "ecologic_objective": ecologic_objective,
        }

        commodity_flows = _solver_operational_flow_records(
            optimization_problem,
            pm_object,
            year,
            country,
            region,
            profile,
        )

        pm_object.set_objective_function_value(
            optimization_problem.objective_function_value
        )
        pm_object.set_instance(optimization_problem.instance)
        _transfer_results_to_parameter_object(
            pm_object,
            optimization_problem.model_type,
        )

        produced_by_commodity = {
            commodity.get_name(): commodity.get_produced_quantity()
            for commodity in pm_object.get_final_commodities_objects()
        }
        demanded_production = {
            commodity.get_name(): commodity.get_produced_quantity()
            for commodity in pm_object.get_final_commodities_objects()
            if commodity.is_demanded()
        }
        produced_quantity = sum(demanded_production.values())
        total_costs = economic_objective
        cost_per_produced_unit = (
            total_costs / produced_quantity
            if total_costs is not None and produced_quantity
            else None
        )

        components = []
        capacity_columns = {}
        final_commodity_names = set(produced_by_commodity)
        for component in pm_object.get_final_components_objects():
            component_name = component.get_name()
            solver_capacity = _solver_nominal_capacity(
                optimization_problem,
                component_name,
            )
            capacity_ratio = (
                component.get_capex_ratio()
                if component.get_component_type() == "conversion"
                else 1.0
            )
            reported_capacity = solver_capacity * capacity_ratio
            transferred_capacity = component.get_fixed_capacity()

            if component.get_component_type() == "conversion":
                produced_by_component = _solver_component_output_quantity(
                    optimization_problem,
                    pm_object,
                    component_name,
                    final_commodity_names,
                )
            elif component.get_component_type() == "generator":
                produced_by_component = _solver_generator_output_quantity(
                    optimization_problem,
                    pm_object,
                    component_name,
                    component.get_generated_commodity(),
                )
            else:
                produced_by_component = 0.0

            if (
                produced_by_component > ZERO_CAPACITY_OUTPUT_SUM_TOLERANCE
                and solver_capacity <= BALANCE_TOLERANCE
            ):
                raise RuntimeError(
                    "Inconsistent optimization result: component "
                    f"'{component_name}' produced {produced_by_component} "
                    "but its solver nominal capacity is zero. "
                    f"Tolerance={ZERO_CAPACITY_OUTPUT_SUM_TOLERANCE}."
                )

            component_row = {
                "scenario_year": year,
                "country": country,
                "region": region,
                "profile": profile,
                "component": component_name,
                "component_type": component.get_component_type(),
                "capacity": reported_capacity,
                "solver_capacity": solver_capacity,
                "transferred_capacity": transferred_capacity,
                "solver_output_quantity": produced_by_component,
                "capacity_basis": (
                    component.get_capex_basis()
                    if component.get_component_type() == "conversion"
                    else "native"
                ),
                "investment": component.get_investment(),
                "annualized_investment": component.get_annualized_investment(),
                "fixed_costs": component.get_total_fixed_costs(),
                "variable_costs": component.get_total_variable_costs(),
                "component_total_costs": component.get_total_costs(),
            }
            components.append(component_row)
            capacity_columns[f"capacity_{component_name}"] = reported_capacity

        commodity_columns = {
            f"produced_{name}": quantity
            for name, quantity in produced_by_commodity.items()
        }
        return {
            "result": {
                **base_result,
                "status": "optimal",
                "total_costs": total_costs,
                "produced_quantity": produced_quantity,
                "cost_per_produced_unit": cost_per_produced_unit,
                **commodity_columns,
                **capacity_columns,
            },
            "components": components,
            "commodity_flows": commodity_flows,
            "error": None,
        }

    except OptimizationNotOptimalError:
        raise
    except ProfileOptimizationError:
        raise
    except Exception as exc:
        if _optimization_profile_retryable(exc) and not job.get(
            "_local_file_retry"
        ):
            safe_country = "".join(
                character if character.isalnum() else "_"
                for character in country
            )
            source_profile = Path(job["profile_dir"]) / profile
            last_retry_error: BaseException = exc

            for attempt in range(1, OPTIMIZATION_PROFILE_RETRIES + 1):
                local_profile_dir = Path(
                    tempfile.mkdtemp(
                        prefix=f"ptx_now_retry_{year}_{safe_country}_",
                    )
                )
                local_profile = local_profile_dir / profile
                try:
                    _copy_profile_with_retries(
                        source=source_profile,
                        destination=local_profile,
                        max_attempts=1,
                    )
                    retry_job = dict(job)
                    retry_job["profile_dir"] = local_profile_dir
                    retry_job["_local_file_retry"] = True
                    return _run_single_profile(retry_job)
                except Exception as retry_exc:  # noqa: BLE001 - flaky profile reads.
                    last_retry_error = retry_exc
                    if (
                        not _optimization_profile_retryable(retry_exc)
                        or attempt >= OPTIMIZATION_PROFILE_RETRIES
                    ):
                        raise ProfileOptimizationError(
                            year=year,
                            country=country,
                            region=region,
                            profile=profile,
                            profile_path=str(local_profile),
                            exception_type=type(retry_exc).__name__,
                            detail=(
                                f"Retry attempt {attempt}/"
                                f"{OPTIMIZATION_PROFILE_RETRIES} failed: "
                                f"{retry_exc!r}"
                            ),
                            traceback_text=traceback.format_exc(),
                        ) from retry_exc
                    time.sleep(SERVER_ACCESS_RETRY_DELAY_SECONDS)
                finally:
                    shutil.rmtree(local_profile_dir, ignore_errors=True)

            raise ProfileOptimizationError(
                year=year,
                country=country,
                region=region,
                profile=profile,
                profile_path=str(source_profile),
                exception_type=type(last_retry_error).__name__,
                detail=repr(last_retry_error),
                traceback_text=traceback.format_exc(),
            ) from last_retry_error

        raise ProfileOptimizationError(
            year=year,
            country=country,
            region=region,
            profile=profile,
            profile_path=str(Path(job["profile_dir"]) / profile),
            exception_type=type(exc).__name__,
            detail=repr(exc),
            traceback_text=traceback.format_exc(),
        ) from exc


def _run_country_jobs(jobs: list[dict[str, Any]], cores: int) -> list[dict[str, Any]]:
    total = len(jobs)
    if total == 0:
        return []
    first_job = jobs[0]
    label = f"{first_job['year']} {first_job['country']} optimization"
    _print_progress(label, 0, total)
    if cores == 1:
        results = []
        for completed, job in enumerate(jobs, start=1):
            results.append(_run_single_profile(job))
            _print_progress(label, completed, total)
        _finish_progress(label, len(results), total)
        return results

    with multiprocessing.Pool(
        processes=cores,
        maxtasksperchild=1,
    ) as pool:
        results = []
        for completed, result in enumerate(
            pool.imap_unordered(_run_single_profile, jobs),
            start=1,
        ):
            results.append(result)
            _print_progress(label, completed, total)
        _finish_progress(label, len(results), total)
        return results


def _safe_excel_value(value: Any) -> Any:
    try:
        import pandas as pd

        return None if pd.isna(value) else value
    except TypeError:
        return value


def _normalize_result_record(record: dict[str, Any]) -> dict[str, Any]:
    normalized_record = {}
    for key, value in record.items():
        normalized_key = key
        if key.startswith("capacity__"):
            normalized_key = "capacity_" + key.removeprefix("capacity__")
        elif key.startswith("produced__"):
            normalized_key = "produced_" + key.removeprefix("produced__")
        normalized_record[normalized_key] = value
    return normalized_record


def _records_from_sheet(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    import pandas as pd

    try:
        frame = pd.read_excel(path, sheet_name=sheet_name)
    except ValueError:
        return []
    if frame.empty:
        return []
    records = []
    for record in frame.to_dict(orient="records"):
        normalized_record = _normalize_result_record(record)
        normalized_record = {
            key: _safe_excel_value(value)
            for key, value in normalized_record.items()
        }
        records.append(normalized_record)
    return records


def _csv_output_path(output_path: Path, sheet_name: str) -> Path:
    return output_path.with_name(f"{output_path.stem}_{sheet_name}.csv")


def _records_from_output(output_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    import pandas as pd

    csv_path = _csv_output_path(output_path, sheet_name)
    if csv_path.is_file():
        frame = pd.read_csv(csv_path)
        if frame.empty:
            return []
        records = []
        for record in frame.to_dict(orient="records"):
            normalized_record = _normalize_result_record(record)
            normalized_record = {
                key: _safe_excel_value(value)
                for key, value in normalized_record.items()
            }
            records.append(normalized_record)
        return records

    if not output_path.is_file():
        return []
    return _records_from_sheet(output_path, sheet_name)


def _load_existing_year_results(
    output_path: Path,
    year: int,
    include_commodity_flows: bool = False,
) -> tuple:
    output_names = [
        "results",
        "components",
        "commodity_flows",
        "parameters_applied",
        PROFILE_FEATURE_SHEET,
        "errors",
    ]
    if not output_path.is_file() and not any(
        _csv_output_path(output_path, sheet_name).is_file()
        for sheet_name in output_names
    ):
        if include_commodity_flows:
            return [], [], [], [], [], []
        return [], [], [], []

    results = _records_from_output(output_path, "results")
    components = _records_from_output(output_path, "components")
    commodity_flows = _records_from_output(output_path, "commodity_flows")
    parameters = _records_from_output(output_path, "parameters_applied")
    profile_features = _records_from_output(output_path, PROFILE_FEATURE_SHEET)
    errors = _records_from_output(output_path, "errors")

    for sheet_name, records in [
        ("results", results),
        ("components", components),
        ("commodity_flows", commodity_flows),
        ("parameters_applied", parameters),
        (PROFILE_FEATURE_SHEET, profile_features),
        ("errors", errors),
    ]:
        wrong_years = {
            int(record.get("scenario_year", record.get("year")))
            for record in records
            if record.get("scenario_year", record.get("year")) is not None
            and int(record.get("scenario_year", record.get("year"))) != year
        }
        if wrong_years:
            raise ValueError(
                f"Existing sheet '{sheet_name}' in {output_path} contains "
                f"unexpected scenario years: {sorted(wrong_years)}"
            )

    if include_commodity_flows:
        return (
            results,
            components,
            commodity_flows,
            parameters,
            profile_features,
            errors,
        )
    return results, components, parameters, errors


def _progress_path(config: RunnerConfig) -> Path:
    return config.output_dir / PROGRESS_FILE_NAME


def _load_progress(config: RunnerConfig) -> dict[str, Any]:
    path = _progress_path(config)
    if not path.is_file():
        return {"version": 1, "scenario_years": {}}

    with path.open("r", encoding="utf-8") as file:
        progress = json.load(file)
    if progress.get("version") != 1:
        raise ValueError(
            f"Unsupported progress file version in {path}: "
            f"{progress.get('version')}"
        )
    progress.setdefault("scenario_years", {})
    return progress


def _save_progress(config: RunnerConfig, progress: dict[str, Any]) -> None:
    path = _progress_path(config)
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = path.with_suffix(".tmp.json")
    with temporary_path.open("w", encoding="utf-8") as file:
        json.dump(progress, file, indent=2, ensure_ascii=True)
    os.replace(temporary_path, path)


def _completed_country_records(
    progress: dict[str, Any],
    year: int,
) -> dict[str, Any]:
    year_state = progress.setdefault("scenario_years", {}).setdefault(
        str(year),
        {},
    )
    return year_state.setdefault("completed_countries", {})


def _mark_country_completed(
    config: RunnerConfig,
    progress: dict[str, Any],
    year: int,
    country: str,
    region: str,
    profile_count: int,
    output_path: Path,
) -> None:
    completed = _completed_country_records(progress, year)
    completed[country] = {
        "region": region,
        "profile_count": profile_count,
        "completed_at": datetime.now().isoformat(timespec="seconds"),
        "output_file": str(output_path),
    }
    _save_progress(config, progress)


def _remove_country_records(
    records: list[dict[str, Any]],
    country: str,
) -> list[dict[str, Any]]:
    return [
        record
        for record in records
        if record.get("country") != country
    ]


def _profile_feature_record_is_complete(record: dict[str, Any]) -> bool:
    status = record.get("profile_feature_status")
    if status is None:
        return record.get("profile_rows") is not None
    return str(status).strip().lower() == "ok"


def _remove_country_profile_records(
    records: list[dict[str, Any]],
    country: str,
    profiles: set[str],
) -> list[dict[str, Any]]:
    return [
        record
        for record in records
        if record.get("country") != country
        or str(record.get("profile")) not in profiles
    ]


def _format_output_workbook(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="center")

        for cell in sheet[1]:
            header = str(cell.value or "").lower()
            if header == "wacc":
                for data_cell in sheet.iter_cols(
                    min_col=cell.column,
                    max_col=cell.column,
                    min_row=2,
                ):
                    for item in data_cell:
                        item.number_format = "0.00%"
            elif (
                "cost" in header
                or "capacity" in header
                or "quantity" in header
                or header in {"economic_objective", "ecologic_objective", "value"}
            ):
                for data_cell in sheet.iter_cols(
                    min_col=cell.column,
                    max_col=cell.column,
                    min_row=2,
                ):
                    for item in data_cell:
                        item.number_format = "#,##0.000000"

        for column_cells in sheet.columns:
            values = [str(cell.value) for cell in column_cells if cell.value is not None]
            width = min(max((len(value) for value in values), default=8) + 2, 42)
            sheet.column_dimensions[column_cells[0].column_letter].width = width

        if sheet.max_row >= 2 and sheet.max_column >= 1:
            table_name = f"tbl_{sheet.title.replace(' ', '_')}"
            table = Table(displayName=table_name[:255], ref=sheet.dimensions)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)

    workbook.save(path)


def write_year_results(
    output_path: Path,
    year: int,
    completed_countries: list[str],
    results: list[dict[str, Any]],
    components: list[dict[str, Any]],
    applied_parameters: list[dict[str, Any]],
    errors: list[dict[str, Any]],
    commodity_flows: list[dict[str, Any]] | None = None,
    profile_features: list[dict[str, Any]] | None = None,
    run_optimization: bool = RUN_OPTIMIZATION,
    read_profile_statistics: bool = READ_PROFILE_STATISTICS,
    csv_write_mode: str = "rewrite",
    append_country: str | None = None,
) -> None:
    import pandas as pd

    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = output_path.with_suffix(".tmp.xlsx")
    results = [_normalize_result_record(record) for record in results]
    if commodity_flows is None:
        commodity_flows = []
    if profile_features is None:
        profile_features = []

    result_base_columns = [
        "scenario_year",
        "country",
        "region",
        "profile",
        "wacc",
        "status",
        "solver_status",
        "economic_objective",
        "ecologic_objective",
        "total_costs",
        "produced_quantity",
        "cost_per_produced_unit",
    ]
    result_extra_columns = sorted(
        {
            key
            for row in results
            for key in row
            if key not in result_base_columns
        }
    )
    results_df = pd.DataFrame(
        results,
        columns=result_base_columns + result_extra_columns,
    )
    components_df = pd.DataFrame(
        components,
        columns=[
            "scenario_year",
            "country",
            "region",
            "profile",
            "component",
            "component_type",
            "capacity",
            "solver_capacity",
            "transferred_capacity",
            "solver_output_quantity",
            "capacity_basis",
            "investment",
            "annualized_investment",
            "fixed_costs",
            "variable_costs",
            "component_total_costs",
        ],
    )
    commodity_flows_df = pd.DataFrame(
        commodity_flows,
        columns=[
            "scenario_year",
            "country",
            "region",
            "profile",
            "commodity",
            "available",
            "purchased",
            "generated",
            "conversion_output",
            "storage_out",
            "emitted",
            "sold",
            "demanded",
            "conversion_input",
            "storage_in",
            "hot_standby_demand",
            "source_total",
            "sink_total",
            "balance_residual",
        ],
    )
    parameters_df = pd.DataFrame(
        applied_parameters,
        columns=[
            "country",
            "region",
            "wacc",
            "year",
            "scope",
            "scope_name",
            "component",
            "parameter",
            "value",
            "unit",
            "source",
            "note",
        ],
    )
    profile_feature_base_columns = [
        "scenario_year",
        "country",
        "region",
        "profile",
        "profile_feature_status",
        "profile_feature_error",
        "profile_rows",
        "total_weighted_hours",
        "weighting_column_present",
    ]
    profile_feature_extra_columns = sorted(
        {
            key
            for row in profile_features
            for key in row
            if key not in profile_feature_base_columns
        }
    )
    profile_features_df = pd.DataFrame(
        profile_features,
        columns=profile_feature_base_columns + profile_feature_extra_columns,
    )
    errors_df = pd.DataFrame(
        errors,
        columns=[
            "scenario_year",
            "country",
            "region",
            "profile",
            "error",
        ],
    )

    summary = pd.DataFrame(
        [
            {"metric": "scenario_year", "value": year},
            {"metric": "run_optimization", "value": run_optimization},
            {
                "metric": "read_profile_statistics",
                "value": read_profile_statistics,
            },
            {"metric": "last_checkpoint", "value": datetime.now().isoformat(timespec="seconds")},
            {"metric": "countries_completed", "value": len(completed_countries)},
            {"metric": "country_names", "value": ", ".join(completed_countries)},
            {"metric": "profile_runs", "value": len(results_df)},
            {
                "metric": "optimal_runs",
                "value": int((results_df.get("status") == "optimal").sum())
                if not results_df.empty
                else 0,
            },
            {
                "metric": "failed_runs",
                "value": int((results_df.get("status") == "failed").sum())
                if not results_df.empty
                else 0,
            },
        ]
    )

    detail_frames = {
        "results": results_df,
        "components": components_df,
        "commodity_flows": commodity_flows_df,
        "parameters_applied": parameters_df,
        PROFILE_FEATURE_SHEET: profile_features_df,
        "errors": errors_df,
    }
    csv_paths = {
        sheet_name: _csv_output_path(output_path, sheet_name)
        for sheet_name in detail_frames
    }
    csv_summary = pd.DataFrame(
        [
            {
                "sheet": sheet_name,
                "rows": len(frame),
                "csv_file": csv_paths[sheet_name].name,
                "included_in_workbook": False,
            }
            for sheet_name, frame in detail_frames.items()
        ]
    )
    summary = pd.concat(
        [
            summary,
            pd.DataFrame(
                [
                    {
                        "metric": f"{sheet_name}_csv",
                        "value": csv_paths[sheet_name].name,
                    }
                    for sheet_name in detail_frames
                ]
            ),
        ],
        ignore_index=True,
    )

    if csv_write_mode not in {"rewrite", "append_country"}:
        raise ValueError(
            "csv_write_mode must be 'rewrite' or 'append_country'."
        )
    if csv_write_mode == "append_country" and not append_country:
        raise ValueError(
            "append_country is required when csv_write_mode='append_country'."
        )

    temporary_csv_paths = {}
    for sheet_name, frame in detail_frames.items():
        csv_path = csv_paths[sheet_name]
        csv_path.parent.mkdir(parents=True, exist_ok=True)
        if csv_write_mode == "append_country":
            if "country" in frame.columns:
                chunk = frame[frame["country"] == append_country]
            else:
                chunk = frame.iloc[0:0]
            csv_exists = csv_path.is_file()
            csv_has_expected_columns = False
            if csv_exists:
                existing_columns = pd.read_csv(csv_path, nrows=0).columns.tolist()
                csv_has_expected_columns = existing_columns == list(frame.columns)
            if (
                (csv_exists and csv_has_expected_columns)
                or (not csv_exists and len(chunk) == len(frame))
            ):
                if chunk.empty:
                    continue
                write_header = (
                    not csv_exists
                    or csv_path.stat().st_size == 0
                )
                chunk.to_csv(
                    csv_path,
                    mode="a",
                    index=False,
                    header=write_header,
                )
                continue

        temporary_csv_path = csv_path.with_suffix(".tmp.csv")
        frame.to_csv(temporary_csv_path, index=False)
        temporary_csv_paths[sheet_name] = temporary_csv_path

    with pd.ExcelWriter(temporary_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="summary", index=False)
        csv_summary.to_excel(writer, sheet_name="csv_outputs", index=False)

    _format_output_workbook(temporary_path)
    for sheet_name, temporary_csv_path in temporary_csv_paths.items():
        os.replace(temporary_csv_path, csv_paths[sheet_name])
    os.replace(temporary_path, output_path)


def _applied_rows_with_country(
    applied: list[dict[str, Any]],
    settings: CountrySettings,
    year: int,
    effective_wacc: float,
) -> list[dict[str, Any]]:
    rows = [
        {
            "country": settings.country,
            "region": settings.region,
            "wacc": effective_wacc,
            **{key: _safe_excel_value(value) for key, value in row.items()},
        }
        for row in applied
    ]
    rows.append(
        {
            "country": settings.country,
            "region": settings.region,
            "wacc": effective_wacc,
            "year": year,
            "scope": "country",
            "scope_name": settings.country,
            "component": "__project__",
            "parameter": "wacc",
            "value": effective_wacc,
            "unit": "fraction",
            "source": settings.wacc_source,
            "note": "Effective WACC applied by runner.",
        }
    )
    return rows


def run(config: RunnerConfig) -> None:
    countries_df, parameters_df = _read_parameter_workbook(
        config.parameters_xlsx
    )
    wacc_overrides = _read_wacc_file(config.wacc_file)
    if wacc_overrides:
        print(
            f"WACC file loaded: {len(wacc_overrides)} country/year "
            "override entries."
        )
    country_dirs = _discover_countries(config)
    if not country_dirs:
        raise SystemExit("No country folders found.")

    base_pm_object = (
        _build_base_parameter_object(config)
        if config.run_optimization
        else None
    )
    progress = _load_progress(config)

    for year in config.scenario_years:
        print(f"\n=== Scenario year {year} ===")
        output_path = config.output_dir / f"country_profile_results_{year}.xlsx"

        (
            year_results,
            year_components,
            year_commodity_flows,
            year_parameters,
            year_profile_features,
            year_errors,
        ) = _load_existing_year_results(
            output_path,
            year,
            include_commodity_flows=True,
        )

        logged_completed = _completed_country_records(progress, year)
        resumable_completed = set()
        for country, record in (
            logged_completed.items()
            if config.run_optimization
            else []
        ):
            country_rows = [
                row
                for row in year_results
                if row.get("country") == country
            ]
            country_component_rows = [
                row
                for row in year_components
                if row.get("country") == country
            ]
            country_flow_rows = [
                row
                for row in year_commodity_flows
                if row.get("country") == country
            ]
            expected_profiles = int(record.get("profile_count", 0))
            has_current_capacity_schema = (
                bool(country_component_rows)
                and all(
                    row.get("solver_capacity") is not None
                    and row.get("transferred_capacity") is not None
                    and row.get("solver_output_quantity") is not None
                    for row in country_component_rows
                )
            )
            has_operational_balance_schema = (
                bool(country_flow_rows)
                and all(
                    row.get("source_total") is not None
                    and row.get("sink_total") is not None
                    and row.get("balance_residual") is not None
                    for row in country_flow_rows
                )
            )
            if (
                country_rows
                and len(country_rows) == expected_profiles
                and has_current_capacity_schema
                and has_operational_balance_schema
                and all(
                    row.get("status") == "optimal"
                    for row in country_rows
                )
            ):
                resumable_completed.add(country)
            else:
                print(
                    f"Progress entry for {year}/{country} is not backed by "
                    "a complete optimal result set using the current output "
                    "schema; country will be rerun."
                )

        completed_countries = sorted(resumable_completed)
        if completed_countries:
            print(
                f"Resume {year}: {len(completed_countries)} completed "
                "countries loaded from progress log."
            )

        for country_dir in country_dirs:
            country = country_dir.name

            if country == "00_Information":
                continue
            if country in resumable_completed:
                expected_profiles = int(
                    logged_completed.get(country, {}).get("profile_count", 0)
                )
                country_feature_rows = [
                    row
                    for row in year_profile_features
                    if row.get("country") == country
                ]
                completed_feature_profiles = {
                    str(row.get("profile"))
                    for row in country_feature_rows
                    if _profile_feature_record_is_complete(row)
                }
                failed_feature_profiles = {
                    str(row.get("profile"))
                    for row in country_feature_rows
                    if not _profile_feature_record_is_complete(row)
                }
                has_failed_feature_rows = any(
                    not _profile_feature_record_is_complete(row)
                    for row in country_feature_rows
                )
                if (
                    config.read_profile_statistics
                    and expected_profiles
                    and (
                        len(completed_feature_profiles) != expected_profiles
                        or has_failed_feature_rows
                    )
                ):
                    print(
                        f"{year}: Backfill profile features for completed "
                        f"country {country}"
                    )
                    progress_record = logged_completed.get(country, {})
                    region = str(progress_record.get("region") or "")
                    if not region:
                        region = _country_settings(
                            countries_df,
                            country,
                            year,
                            wacc_overrides,
                        ).region
                    profile_dir = _profile_dir(config, country_dir, year)
                    _wait_for_profile_dir(profile_dir)
                    profiles = _profile_files(
                        profile_dir,
                        config.recursive_profiles,
                    )
                    profiles_to_retry = [
                        profile
                        for profile in profiles
                        if profile not in completed_feature_profiles
                        or profile in failed_feature_profiles
                    ]
                    profiles_to_retry_set = set(profiles_to_retry)
                    if profiles_to_retry:
                        print(
                            f"{year}: Retry {len(profiles_to_retry)} profile "
                            f"feature row(s) for completed country {country}"
                        )
                    year_profile_features = _remove_country_profile_records(
                        year_profile_features,
                        country,
                        profiles_to_retry_set,
                    )
                    if profiles_to_retry:
                        year_profile_features.extend(
                            _profile_feature_records(
                                profile_dir,
                                profiles_to_retry,
                                year=year,
                                country=country,
                                region=region,
                                workers=config.profile_statistics_cores,
                            )
                        )
                    write_year_results(
                        output_path=output_path,
                        year=year,
                        completed_countries=completed_countries,
                        results=year_results,
                        components=year_components,
                        commodity_flows=year_commodity_flows,
                        applied_parameters=year_parameters,
                        profile_features=year_profile_features,
                        errors=year_errors,
                        run_optimization=config.run_optimization,
                        read_profile_statistics=config.read_profile_statistics,
                    )
                print(f"{year}: Skip completed country {country}")
                continue

            print(f"{year}: {country}")
            country_successful = False
            completed_profile_count = 0
            country_had_existing_rows = any(
                row.get("country") == country
                for rows in [
                    year_results,
                    year_components,
                    year_commodity_flows,
                    year_parameters,
                    year_profile_features,
                    year_errors,
                ]
                for row in rows
            )

            # Remove stale rows only for the output branches that are being
            # recalculated. A statistics-only run must preserve optimization
            # results that may already exist in the workbook.
            if config.run_optimization:
                year_results = _remove_country_records(year_results, country)
                year_components = _remove_country_records(
                    year_components,
                    country,
                )
                year_commodity_flows = _remove_country_records(
                    year_commodity_flows,
                    country,
                )
                year_parameters = _remove_country_records(
                    year_parameters,
                    country,
                )
                year_errors = _remove_country_records(year_errors, country)
            if config.read_profile_statistics:
                year_profile_features = _remove_country_records(
                    year_profile_features,
                    country,
                )

            try:
                settings = _country_settings(
                    countries_df,
                    country,
                    year,
                    wacc_overrides,
                )
                profile_dir = _profile_dir(
                    config,
                    country_dir,
                    year,
                )
                _wait_for_profile_dir(profile_dir)
                profiles = _profile_files(
                    profile_dir,
                    config.recursive_profiles,
                )
                if not profiles:
                    raise FileNotFoundError(
                        f"No profile files found in {profile_dir}"
                    )

                if config.read_profile_statistics:
                    country_profile_features = _profile_feature_records(
                        profile_dir,
                        profiles,
                        year=year,
                        country=country,
                        region=settings.region,
                        workers=config.profile_statistics_cores,
                    )
                    year_profile_features.extend(country_profile_features)

                if config.run_optimization:
                    parameter_rows = _parameter_rows_for_country(
                        parameters_df,
                        year,
                        settings,
                    )
                    if parameter_rows.empty:
                        raise ValueError(
                            f"No active parameters found for {country}, "
                            f"region {settings.region}, year {year}."
                        )

                    country_pm_object = deepcopy(base_pm_object)
                    applied = _apply_parameters(
                        country_pm_object,
                        parameter_rows,
                    )
                    if settings.wacc is not None:
                        country_pm_object.set_wacc(settings.wacc)
                    _print_country_assumptions(
                        year,
                        settings,
                        parameter_rows,
                        country_pm_object,
                    )

                    jobs = [
                        {
                            "pm_object": deepcopy(country_pm_object),
                            "profile_dir": profile_dir,
                            "profile": profile,
                            "country": country,
                            "region": settings.region,
                            "year": year,
                            "wacc": country_pm_object.get_wacc(),
                            "solver": config.solver,
                        }
                        for profile in profiles
                    ]
                    country_results = _run_country_jobs(jobs, config.cores)

                    unsuccessful = [
                        result
                        for result in country_results
                        if result.get("result", {}).get("status") != "optimal"
                    ]
                    if unsuccessful:
                        failed_result = unsuccessful[0]
                        result_row = failed_result.get("result", {})
                        error_row = failed_result.get("error") or {}
                        raise ProfileOptimizationError(
                            year=year,
                            country=country,
                            region=settings.region,
                            profile=str(result_row.get("profile", "unknown")),
                            profile_path=str(
                                profile_dir
                                / str(result_row.get("profile", "unknown"))
                            ),
                            exception_type=str(
                                result_row.get("solver_status", "unknown")
                            ),
                            detail=str(
                                error_row.get(
                                    "error",
                                    f"Non-optimal result row: {result_row}",
                                )
                            ),
                            traceback_text="No worker traceback was returned.",
                        )

                    year_results.extend(
                        result["result"] for result in country_results
                    )
                    year_components.extend(
                        row
                        for result in country_results
                        for row in result["components"]
                    )
                    year_commodity_flows.extend(
                        row
                        for result in country_results
                        for row in result["commodity_flows"]
                    )
                    year_errors.extend(
                        result["error"]
                        for result in country_results
                        if result["error"] is not None
                    )
                    year_parameters.extend(
                        _applied_rows_with_country(
                            applied,
                            settings,
                            year,
                            country_pm_object.get_wacc(),
                        )
                    )
                    completed_countries.append(country)
                    country_successful = True
                    completed_profile_count = len(country_results)
                else:
                    completed_countries.append(country)
                    completed_profile_count = len(profiles)

            except OptimizationNotOptimalError as exc:
                year_errors.append(
                    {
                        "scenario_year": exc.year,
                        "country": exc.country,
                        "region": settings.region,
                        "profile": exc.profile,
                        "error": str(exc),
                    }
                )
                write_year_results(
                    output_path=output_path,
                    year=year,
                    completed_countries=completed_countries,
                    results=year_results,
                    components=year_components,
                    commodity_flows=year_commodity_flows,
                    applied_parameters=year_parameters,
                    profile_features=year_profile_features,
                    errors=year_errors,
                    run_optimization=config.run_optimization,
                    read_profile_statistics=config.read_profile_statistics,
                )
                print(f"\nABORT: {exc}")
                print(f"Checkpoint written before abort: {output_path}")
                raise SystemExit(str(exc)) from exc

            except ProfileOptimizationError as exc:
                year_errors.append(
                    {
                        "scenario_year": exc.year,
                        "country": exc.country,
                        "region": exc.region,
                        "profile": exc.profile,
                        "error": str(exc),
                    }
                )
                write_year_results(
                    output_path=output_path,
                    year=year,
                    completed_countries=completed_countries,
                    results=year_results,
                    components=year_components,
                    commodity_flows=year_commodity_flows,
                    applied_parameters=year_parameters,
                    profile_features=year_profile_features,
                    errors=year_errors,
                    run_optimization=config.run_optimization,
                    read_profile_statistics=config.read_profile_statistics,
                )
                print(f"\nABORT: {exc}")
                print(f"Checkpoint written before abort: {output_path}")
                raise SystemExit(str(exc)) from exc

            except Exception as exc:  # noqa: BLE001 - checkpoint and continue.
                year_errors.append(
                    {
                        "scenario_year": year,
                        "country": country,
                        "region": None,
                        "profile": None,
                        "error": repr(exc),
                    }
                )
                print(f"Skip {country}: {exc}")

            write_year_results(
                output_path=output_path,
                year=year,
                completed_countries=completed_countries,
                results=year_results,
                components=year_components,
                commodity_flows=year_commodity_flows,
                applied_parameters=year_parameters,
                profile_features=year_profile_features,
                errors=year_errors,
                run_optimization=config.run_optimization,
                read_profile_statistics=config.read_profile_statistics,
                csv_write_mode=(
                    "rewrite"
                    if country_had_existing_rows
                    else "append_country"
                ),
                append_country=country,
            )
            print(f"Checkpoint written: {output_path}")
            if country_successful:
                _mark_country_completed(
                    config=config,
                    progress=progress,
                    year=year,
                    country=country,
                    region=settings.region,
                    profile_count=completed_profile_count,
                    output_path=output_path,
                )
                resumable_completed.add(country)
                print(
                    f"Progress logged: year={year}, country={country}, "
                    f"profiles={completed_profile_count}"
                )

        print(f"Final result for {year}: {output_path}")


def main() -> None:
    config = build_config()
    validate_config(config)
    print(
        f"Runner version: {RUNNER_VERSION}\n"
        f"Countries: {config.countries_root}\n"
        f"YAML: {config.settings_yaml}\n"
        f"Parameters: {config.parameters_xlsx}\n"
        f"WACC file: {config.wacc_file}\n"
        f"Years: {config.scenario_years}\n"
        f"Run optimization: {config.run_optimization}\n"
        f"Read profile statistics: {config.read_profile_statistics}\n"
        f"Profile statistics workers: {config.profile_statistics_cores}\n"
        f"Profile statistics retries: {PROFILE_STATISTICS_RETRIES}\n"
        f"Optimization profile retries: {OPTIMIZATION_PROFILE_RETRIES}\n"
        f"Workers per country: {config.cores}"
    )
    run(config)


if __name__ == "__main__":
    multiprocessing.freeze_support()
    main()
